#! /usr/bin/python3
# Python program to parse header file and read all prototypes of function and insert it into excel sheet with unique ID start with IDX0
import openpyxl
import re

def parse(file_parse):
    fd = open(file_parse, 'r')
    parse = fd.readlines()
    fd.close()
    data = []
    for i in parse:
        if "(" in i and ")" in i and ";" in i:  #to not count the struct , enum , comments , macros , etc.
            data.append(i)
        else:
            pass
    for i in range(len(data)):
        data[i] = data[i].replace('\n', '')

    return data


def read(file_generate):
    Workbook = openpyxl.load_workbook(file_generate)
    sheet = Workbook.active

    for raw in sheet.iter_rows(values_only = True):
        print(raw)
    Workbook.close()

def write(file_generate,content):
    Workbook = openpyxl.Workbook()
    sheet = Workbook.active
    sheet['A1'] = "Prototype"
    sheet['B1'] = "ID"
    sheet['C1'] = "return_type"
    sheet['D1'] = "parameters"
    number = 0
    for i in content:
        #prototype
        tmp = [i]
        #ID
        tmp.append(f"IDX{number}")
        #return type
        tmp.append(i.split()[0])
        #parameters
        match = re.search(r'\((.*?)\)', i)
        tmp.append(match.group(1))
        number +=1
        sheet.append(tmp)
        
    Workbook.save(file_generate)
    Workbook.close()


write('function_prototypes.xlsx',parse('header.h'))
read('function_prototypes.xlsx')
 

 
#================================================================
fd = open('header.h', 'r')
parse = fd.read()
fd.close()
pattern = r'\b\w+\s+(\w+)\s*\([^)]*\)\s*;'
matches = re.findall(pattern, parse)
print(matches) #['LCD_init', 'LCD_send_command', 'LCD_send_char', 'LCD_send_string', 'Display_number', 'Display_float']
'''
\b asserts the start of the word.
\w+ captures the return type.
\s+ allows for any amount of whitespace.
(\w+) captures the function name.
\s* allows for any amount of whitespace.
\( matches the opening parenthesis.
[^)]* matches the content between the parentheses (the function arguments).
\) matches the closing parenthesis.
\s* allows for any amount of whitespace.
; matches the semicolon at the end of the function prototype.
'''
