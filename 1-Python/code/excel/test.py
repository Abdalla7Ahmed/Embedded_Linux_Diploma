#! /usr/bin/python3


import openpyxl

def read(file_path):
    Workbook = openpyxl.load_workbook(file_path)
    sheet = Workbook.active

    for raw in sheet.iter_rows(values_only = True):
        print(raw)
    Workbook.close()

def write(file_path):
    Workbook = openpyxl.Workbook()
    sheet = Workbook.active
    data = [
        ['Name','Age','Gender','Country'],
        ['Abdallah','23','Male','Egypt'],
        ['Ali','15','Male','Egypt'],
        ['Omar','20','Male','Egypt'],
        ['Mai','23','Female','Egypt'],
    ]
    for raw in data:
        sheet.append(raw)

    Workbook.save(file_path)
    Workbook.close()
write('data.xlsx')
read('data.xlsx')

#generate sheet 
'''
Workbook1 = openpyxl.Workbook()
new_sheet = Workbook1.create_sheet(title = "new_sheet")
new_sheet['A1'] = "A1"
new_sheet['B1'] = "B1"
new_sheet['A2'] = "A2"
new_sheet['B2'] = "B2"

Workbook1.save('data.xlsx')

'''

